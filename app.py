"""
排课求解器 Web UI
基于 Streamlit 框架
"""

import streamlit as st
import pandas as pd
import re
from pathlib import Path
import io
import time
import threading
try:
    from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx
except ImportError:
    from streamlit.scriptrunner import add_script_run_ctx, get_script_run_ctx
from ortools.sat.python import cp_model
from collections import defaultdict
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="智能排课求解器",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        text-align: center;
        color: #1f77b4;
        padding: 1rem 0;
    }
    .sub-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #ff7f0e;
        margin-top: 1.5rem;
        margin-bottom: 0.5rem;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        border-left: 5px solid #28a745;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        background-color: #fff3cd;
        border-left: 5px solid #ffc107;
        margin: 1rem 0;
    }
    .error-box {
        padding: 1rem;
        background-color: #f8d7da;
        border-left: 5px solid #dc3545;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        background-color: #d1ecf1;
        border-left: 5px solid #17a2b8;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

def natural_sort_key(s):
    """自然排序的key函数，用于正确排序包含数字的字符串
    例如: S1, S2, S3, ..., S9, S10, S11 (而不是 S1, S10, S11, S2)
    """
    import re
    return [int(text) if text.isdigit() else text.lower() 
            for text in re.split(r'(\d+)', str(s))]

def parse_subject_string(subject_str):
    """解析科目字符串（支持中英文括号及空格）"""
    subjects = {}
    # 增加 \s* 允许括号周围有空格
    pattern = r'([^,\(（]+)\s*[\(（]\s*(\d+)\s*[\)）]'
    matches = re.findall(pattern, subject_str)
    for subject, hours in matches:
        subject = subject.strip()
        subjects[subject] = int(hours)
    return subjects

# 时区函数
def get_malaysia_time():
    """获取马来西亚时间（UTC+8）"""
    import datetime
    # 获取UTC时间
    utc_now = datetime.datetime.utcnow()
    # 转换为马来西亚时间（UTC+8）
    malaysia_time = utc_now + datetime.timedelta(hours=8)
    return malaysia_time

# 存储功能相关函数
SAVED_SOLUTIONS_FILE = "saved_solutions.pkl"

def load_saved_solutions_from_disk():
    """从本地磁盘加载已保存的方案"""
    if not os.path.exists(SAVED_SOLUTIONS_FILE):
        return {}
    try:
        with open(SAVED_SOLUTIONS_FILE, 'rb') as f:
            return pickle.load(f)
    except:
        return {}

def save_saved_solutions_to_disk(saved_solutions):
    """将已保存的方案写入本地磁盘"""
    try:
        with open(SAVED_SOLUTIONS_FILE, 'wb') as f:
            pickle.dump(saved_solutions, f)
        return True
    except Exception as e:
        print(f"❌ 保存方案到磁盘失败: {e}")
        return False

def save_solution_to_storage(sol, save_name):
    """保存方案到存储"""
    timestamp = get_malaysia_time().strftime("%Y-%m-%d %H:%M:%S")
    
    # 保存到 session_state
    st.session_state['saved_solutions'][save_name] = {
        'solution': sol,
        'timestamp': timestamp,
        'original_name': sol['name']
    }
    
    # 同步到磁盘
    if save_saved_solutions_to_disk(st.session_state['saved_solutions']):
        st.toast(f"💾 方案已保存到本地", icon="✅")
    else:
        st.toast(f"⚠️ 保存到磁盘失败", icon="❌")

def delete_saved_solution(save_name):
    """删除已保存的方案"""
    if save_name in st.session_state['saved_solutions']:
        del st.session_state['saved_solutions'][save_name]
        # 同步到磁盘
        save_saved_solutions_to_disk(st.session_state['saved_solutions'])

# Read Excel File
def parse_uploaded_file(uploaded_file):
    """解析上传的Excel/CSV文件"""
    try:
        if uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls'):
            df = pd.read_excel(uploaded_file)
        else:
            encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'big5', 'cp936', 'utf-8-sig']
            df = None
            last_error = None
            
            for encoding in encodings:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, encoding=encoding)
                    st.success(f"✅ 成功读取文件（编码：{encoding}）")
                    break
                except (UnicodeDecodeError, Exception) as e:
                    last_error = e
                    continue
            
            if df is None:
                raise Exception(f"无法识别文件编码，请确保文件是有效的CSV格式。最后错误：{last_error}")
        
        packages = {}
        subject_hours = {}
        total_hours_stats = []
        
        for _, row in df.iterrows():
            package_name = str(row['配套']).strip()
            student_count = int(row['人数'])
            subject_str = str(row['科目'])
            
            subjects = parse_subject_string(subject_str)
            
            total_hours = sum(subjects.values())
            total_hours_stats.append({
                '配套': package_name,
                '总课时': total_hours
            })
            
            packages[package_name] = {
                '人数': student_count,
                '科目': subjects
            }
            
            for subject, hours in subjects.items():
                if subject not in subject_hours:
                    subject_hours[subject] = hours
                elif subject_hours[subject] != hours:
                    st.error(f"❌ **数据错误：科目'{subject}'的课时不一致！**")
                    st.error(f"   • 在某些配套中是 **{subject_hours[subject]}小时**")
                    st.error(f"   • 在'{package_name}'配套中是 **{hours}小时**")
                    st.markdown("---")
                    st.markdown("""
                    ### 🔍 为什么会导致错误？
                    
                    系统会为每个科目创建**统一长度**的班级（如6小时的会计班）。
                    所有学生都会被分配到这些统一的班级中。
                    
                    如果配套A需要6小时会计，配套B需要4小时会计：
                    - ❌ 无法用6小时的班满足4小时的需求
                    - ❌ 也无法用4小时的班满足6小时的需求
                    - ❌ 导致求解器找不到可行解
                    
                    ### ✅ 解决方案：
                    
                    **方案1：统一课时（推荐）**
                    - 将所有配套的'{subject}'课时改为相同值（如都改为6小时或都改为4小时）
                    
                    **方案2：分离科目**
                    - 将4小时的会计命名为"会计1"
                    - 将6小时的会计命名为"会计2"
                    - 这样系统会将它们视为不同科目
                    """)
                    return None, None, None
        
        min_hours = min(s['总课时'] for s in total_hours_stats)
        max_hours = max(s['总课时'] for s in total_hours_stats)
        
        if min_hours < 21:
            st.info(f"ℹ️ 检测到部分配套总课时少于21小时（范围：{min_hours}-{max_hours}小时）")
            st.success("✅ 系统支持总课时不足的配套，这些配套将在某些时段不上课")
            
            short_packages = [s for s in total_hours_stats if s['总课时'] < 21]
            if short_packages:
                with st.expander("查看总课时不足21的配套"):
                    for pkg in short_packages:
                        st.text(f"  {pkg['配套']}: {pkg['总课时']}小时")
        
        return packages, subject_hours, max_hours
    
    except Exception as e:
        st.error(f"❌ 文件解析失败: {str(e)}")
        return None, None, None

def calculate_subject_enrollment(packages):
    """计算每个科目的总选修人数"""
    enrollment = defaultdict(int)
    for p_data in packages.values():
        for subject in p_data['科目'].keys():
            enrollment[subject] += p_data['人数']
    return dict(enrollment)

def calculate_recommended_slots(max_total_hours):
    """根据最大总课时计算推荐的时段组数
    
    时段组结构：前(n-1)个时段组各2小时，最后1个时段组3小时
    总容量 = (n-1)*2 + 3 = 2n+1 小时
    
    参数:
        max_total_hours: 所有配套中的最大总课时
    
    返回:
        推荐的时段组数
    """
    import math
    if max_total_hours <= 3:
        return 1

    recommended = math.ceil((max_total_hours - 1) / 2)
    return max(2, min(recommended, 20))
#Main Algorithms
class ScheduleSolver:
    def __init__(self, packages, subject_hours, config):
        self.packages = packages
        self.subject_hours = subject_hours
        self.config = config
        self.subjects = list(subject_hours.keys())
        self.package_names = list(packages.keys())
        
        # 时段定义
        self.TIME_SLOTS_1H = list(range(1, config['num_slots'] * 2 + 2))
        self.SLOT_GROUPS = {}
        for i in range(1, config['num_slots'] + 1):
            if i < config['num_slots']:
                self.SLOT_GROUPS[f'S{i}'] = [i*2-1, i*2]
            else:
                self.SLOT_GROUPS[f'S{i}'] = [i*2-1, i*2, i*2+1]
        
        self.SLOT_TO_GROUP = {}
        for group_name, slots in self.SLOT_GROUPS.items():
            for slot in slots:
                self.SLOT_TO_GROUP[slot] = group_name
        
        self.subject_enrollment = calculate_subject_enrollment(packages)
    
    def build_model(self, objective_type='min_classes'):
        """构建模型"""
        model = cp_model.CpModel()
        
        u_r = {}
        y_rt = {}
        u_pkr = {}
        x_prt = {}
        
        for k in self.subjects:
            for r in range(1, self.config['max_classes_per_subject'] + 1):
                u_r[(k, r)] = model.NewBoolVar(f'u_{k}_{r}')
                for t in self.TIME_SLOTS_1H:
                    y_rt[(k, r, t)] = model.NewBoolVar(f'y_{k}_{r}_{t}')
        
        for p in self.package_names:
            for k in self.subjects:
                for r in range(1, self.config['max_classes_per_subject'] + 1):
                    u_pkr[(p, k, r)] = model.NewBoolVar(f'u_{p}_{k}_{r}')
                    for t in self.TIME_SLOTS_1H:
                        x_prt[(p, k, r, t)] = model.NewBoolVar(f'x_{p}_{k}_{r}_{t}')

        for k in self.subjects:
            for r in range(2, self.config['max_classes_per_subject'] + 1):
                model.Add(u_r[(k, r)] <= u_r[(k, r - 1)])

        for k in self.subjects:
            for r in range(2, self.config['max_classes_per_subject'] + 1):
                # 计算第 r 班的人数
                size_curr = sum(
                    self.packages[p]['人数'] * u_pkr[(p, k, r)] 
                    for p in self.package_names
                )
                # 计算第 r-1 班的人数
                size_prev = sum(
                    self.packages[p]['人数'] * u_pkr[(p, k, r - 1)] 
                    for p in self.package_names
                )
                
                # 添加约束：后一个班的人数必须小于等于前一个班
                model.Add(size_curr <= size_prev)
                
        # ==============================================================================
        for k in self.subjects:
            H_k = self.subject_hours[k]
            for r in range(1, self.config['max_classes_per_subject'] + 1):
                total_hours = sum(y_rt[(k, r, t)] for t in self.TIME_SLOTS_1H)
                model.Add(total_hours == H_k).OnlyEnforceIf(u_r[(k, r)])
                model.Add(total_hours == 0).OnlyEnforceIf(u_r[(k, r)].Not())
        
        for p in self.package_names:
            for k in self.subjects:
                if k in self.packages[p]['科目']:
                    model.Add(sum(u_pkr[(p, k, r)] for r in range(1, self.config['max_classes_per_subject'] + 1)) == 1)
                else:
                    for r in range(1, self.config['max_classes_per_subject'] + 1):
                        model.Add(u_pkr[(p, k, r)] == 0)
        
        for k in self.subjects:
            for r in range(1, self.config['max_classes_per_subject'] + 1):
                class_size = sum(self.packages[p]['人数'] * u_pkr[(p, k, r)] for p in self.package_names)
                model.Add(class_size >= self.config['min_class_size']).OnlyEnforceIf(u_r[(k, r)])
                model.Add(class_size <= self.config['max_class_size']).OnlyEnforceIf(u_r[(k, r)])
                model.Add(class_size == 0).OnlyEnforceIf(u_r[(k, r)].Not())
        
        for p in self.package_names:
            for k in self.subjects:
                for r in range(1, self.config['max_classes_per_subject'] + 1):
                    for t in self.TIME_SLOTS_1H:
                        model.Add(x_prt[(p, k, r, t)] <= u_pkr[(p, k, r)])
                        model.Add(x_prt[(p, k, r, t)] <= y_rt[(k, r, t)])
                        model.Add(x_prt[(p, k, r, t)] >= u_pkr[(p, k, r)] + y_rt[(k, r, t)] - 1)
        
        for p in self.package_names:
            for t in self.TIME_SLOTS_1H:
                model.Add(sum(x_prt[(p, k, r, t)] 
                            for k in self.subjects 
                            for r in range(1, self.config['max_classes_per_subject'] + 1)) <= 1)
        

        for k in self.subjects:
            for t in self.TIME_SLOTS_1H:
                limit = 1
                
                if self.config.get('enable_concurrency', False):
                    limit = self.config['max_classes_per_subject']
                
                model.Add(sum(y_rt[(k, r, t)] for r in range(1, self.config['max_classes_per_subject'] + 1)) <= limit)
        
        for p in self.package_names:
            for k in self.subjects:
                if k in self.packages[p]['科目']:
                    required_hours = self.packages[p]['科目'][k]
                    total_hours_pk = sum(
                        x_prt[(p, k, r, t)]
                        for r in range(1, self.config['max_classes_per_subject'] + 1)
                        for t in self.TIME_SLOTS_1H
                    )
                    model.Add(total_hours_pk == required_hours)
        
        for k in self.subjects:
            model.Add(sum(u_r[(k, r)] for r in range(1, self.config['max_classes_per_subject'] + 1)) <= self.config['max_classes_per_subject'])
        
        for k, count in self.config['forced_class_count'].items():
            if k in self.subjects:
                model.Add(sum(u_r[(k, r)] for r in range(1, self.config['max_classes_per_subject'] + 1)) == count)
        
        slot_split_penalty = 0
        
        if not self.config['allow_slot_split']:
            for p in self.package_names:
                for group_name, group_slots in self.SLOT_GROUPS.items():
                    subjects_in_group = []
                    for k in self.subjects:
                        for r in range(1, self.config['max_classes_per_subject'] + 1):
                            has_subject = model.NewBoolVar(f'has_{p}_{k}_{r}_{group_name}')
                            model.AddMaxEquality(has_subject, [x_prt[(p, k, r, t)] for t in group_slots])
                            subjects_in_group.append(has_subject)
                    
                    model.Add(sum(subjects_in_group) <= 1)
        
        else:
            split_vars = []
            for p in self.package_names:
                for group_name, group_slots in self.SLOT_GROUPS.items():
                    subjects_in_group = []
                    for k in self.subjects:
                        for r in range(1, self.config['max_classes_per_subject'] + 1):
                            has_subject = model.NewBoolVar(f'has_{p}_{k}_{r}_{group_name}')
                            model.AddMaxEquality(has_subject, [x_prt[(p, k, r, t)] for t in group_slots])
                            subjects_in_group.append(has_subject)
                    
                    num_subjects = sum(subjects_in_group)
                    is_split = model.NewBoolVar(f'split_{p}_{group_name}')
                    model.Add(num_subjects >= 2).OnlyEnforceIf(is_split)
                    model.Add(num_subjects <= 1).OnlyEnforceIf(is_split.Not())
                    split_vars.append(is_split)
            
            slot_split_penalty = sum(split_vars) * self.config['slot_split_penalty']
        
        total_classes = sum(u_r[(k, r)] for k in self.subjects for r in range(1, self.config['max_classes_per_subject'] + 1))
        priority_penalty = sum(
            u_r[(k, r)] * r * max(0, 100 - self.subject_enrollment[k])
            for k in self.subjects 
            for r in range(1, self.config['max_classes_per_subject'] + 1)
        )
        
        if objective_type == 'min_classes':
            model.Minimize(total_classes * 100000 + slot_split_penalty + priority_penalty)
            
        elif objective_type == 'balanced':
            effective_sizes_for_max = []
            effective_sizes_for_min = []
            
            for k in self.subjects:
                for r in range(1, self.config['max_classes_per_subject'] + 1):
                    actual_size = sum(self.packages[p]['人数'] * u_pkr[(p, k, r)] for p in self.package_names)
                    
                    eff_size_max = model.NewIntVar(0, 200, f'eff_max_{k}_{r}')
                    model.Add(eff_size_max == actual_size).OnlyEnforceIf(u_r[(k, r)])
                    model.Add(eff_size_max == 0).OnlyEnforceIf(u_r[(k, r)].Not())
                    effective_sizes_for_max.append(eff_size_max)
                    
                    eff_size_min = model.NewIntVar(0, 200, f'eff_min_{k}_{r}')
                    model.Add(eff_size_min == actual_size).OnlyEnforceIf(u_r[(k, r)])
                    model.Add(eff_size_min == 200).OnlyEnforceIf(u_r[(k, r)].Not())
                    effective_sizes_for_min.append(eff_size_min)
            
            max_size = model.NewIntVar(0, 200, 'max_size')
            min_size = model.NewIntVar(0, 200, 'min_size')
            
            model.AddMaxEquality(max_size, effective_sizes_for_max)
            model.AddMinEquality(min_size, effective_sizes_for_min)


            weight_class = -100 
            
            weight_balance = 200 
            weight_split = self.config.get('slot_split_penalty', 1000) 
            
            model.Minimize(
                total_classes * weight_class + 
                (max_size - min_size) * weight_balance + 
                slot_split_penalty * (weight_split / 100) + 
                priority_penalty
            )

        elif objective_type == 'subject_balanced':
            import math 
            
            total_excess_penalty = 0 
            total_raw_penalty = 0    
            
            allowed_gap = 6      # 允许误差
            scheme_c_max_size = self.config.get('dynamic_max_limit', 30)


            for k in self.subjects:
                # 1. 计算理论硬锁定数
                total_k_students = self.subject_enrollment.get(k, 0)
                if total_k_students > 0:
                    locked_class_count = math.ceil(total_k_students / scheme_c_max_size)
                else:
                    locked_class_count = 0
                
                active_classes_var = sum(u_r[(k, r)] for r in range(1, self.config['max_classes_per_subject'] + 1))
                if not self.config.get('relax_hard_lock', False):
                    model.Add(active_classes_var <= locked_class_count)
                

                k_effective_sizes_max = [] 
                k_effective_sizes_min = [] 
                
                subject_active = model.NewBoolVar(f'active_subj_{k}')
                model.Add(active_classes_var >= 1).OnlyEnforceIf(subject_active)
                model.Add(active_classes_var == 0).OnlyEnforceIf(subject_active.Not())

                for r in range(1, self.config['max_classes_per_subject'] + 1):
                    # 计算班级 r 的实际人数
                    actual_size = sum(
                        self.packages[p]['人数'] * u_pkr[(p, k, r)] 
                        for p in self.package_names
                    )
                    
                    # 强制单班上限
                    model.Add(actual_size <= scheme_c_max_size)

                    # Max/Min 辅助计算
                    eff_max = model.NewIntVar(0, 200, f'eff_max_C_{k}_{r}')
                    model.Add(eff_max == actual_size).OnlyEnforceIf(u_r[(k, r)])
                    model.Add(eff_max == 0).OnlyEnforceIf(u_r[(k, r)].Not())
                    k_effective_sizes_max.append(eff_max)
                    
                    eff_min = model.NewIntVar(0, 200, f'eff_min_C_{k}_{r}')
                    model.Add(eff_min == actual_size).OnlyEnforceIf(u_r[(k, r)])
                    model.Add(eff_min == 200).OnlyEnforceIf(u_r[(k, r)].Not())
                    k_effective_sizes_min.append(eff_min)
                
                # 极差计算
                k_max_size = model.NewIntVar(0, 200, f'k_max_C_{k}')
                k_min_size = model.NewIntVar(0, 200, f'k_min_C_{k}')
                model.AddMaxEquality(k_max_size, k_effective_sizes_max)
                model.AddMinEquality(k_min_size, k_effective_sizes_min)
                
                k_range = model.NewIntVar(0, 200, f'range_C_{k}')
                model.Add(k_range == k_max_size - k_min_size).OnlyEnforceIf(subject_active)
                model.Add(k_range == 0).OnlyEnforceIf(subject_active.Not())
                
                # 误差计算
                k_excess = model.NewIntVar(0, 200, f'excess_C_{k}')
                model.Add(k_excess >= k_range - allowed_gap).OnlyEnforceIf(subject_active)
                model.Add(k_excess >= 0)
                
                total_excess_penalty += k_excess
                total_raw_penalty += k_range

            weight_class_penalty = 0      # 班数已锁死，无需惩罚
            weight_excess = 1000000       # 严禁误差超标
            weight_raw = 100              # 尽量平均
            
            weight_split = self.config.get('slot_split_penalty', 1000)
            
            model.Minimize(
                total_excess_penalty * weight_excess + 
                total_raw_penalty * weight_raw + 
                slot_split_penalty * (weight_split / 100) + 
                priority_penalty
            )
        return model, {'u_r': u_r, 'y_rt': y_rt, 'u_pkr': u_pkr, 'x_prt': x_prt}
    
    class SolutionPrinter(cp_model.CpSolverSolutionCallback):
        def __init__(self, status_placeholder, scheme_name):
            cp_model.CpSolverSolutionCallback.__init__(self)
            self.status_placeholder = status_placeholder
            self.scheme_name = scheme_name
            self.solution_count = 0
            self.start_time = time.time()
            
            try:
                self.ctx = get_script_run_ctx()
            except Exception:
                self.ctx = None

        def on_solution_callback(self):
            if self.ctx:
                add_script_run_ctx(threading.current_thread(), self.ctx)
                
            self.solution_count += 1
            current_time = time.time()
            elapsed = current_time - self.start_time
            
            self.status_placeholder.markdown(
                f"⚙️ **{self.scheme_name}** - 正在疯狂计算... "
                f"(已发现 **{self.solution_count}** 个可行方案, "
                f"耗时: {elapsed:.1f}s)"
            )

    def solve(self, model, variables, timeout, status_placeholder=None, scheme_name=""):
        """求解模型 (优化版：带上下文修复)"""
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = timeout
        solver.parameters.log_search_progress = False
        solver.parameters.num_search_workers = 8
        
        callback = None
        if status_placeholder and scheme_name:
            callback = self.SolutionPrinter(status_placeholder, scheme_name)
        
        start_time = time.time()
        
        if callback:
            status = solver.Solve(model, callback)
        else:
            status = solver.Solve(model)
            
        solve_time = time.time() - start_time
        
        status_map = {
            cp_model.OPTIMAL: ('最优解', '✅'),
            cp_model.FEASIBLE: ('可行解', '✅'),
            cp_model.INFEASIBLE: ('无解', '❌'),
            cp_model.MODEL_INVALID: ('模型无效', '⚠️'),
            cp_model.UNKNOWN: ('超时/未知', '⏱️')
        }
        
        status_name, icon = status_map.get(status, ('未知状态', '❓'))
        
        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            return {
                'status': 'success',
                'solver': solver,
                'variables': variables,
                'solve_status': status_name,
                'icon': icon,
                'solve_time': solve_time
            }
        else:
            return {
                'status': 'failed',
                'solve_status': status_name,
                'icon': icon,
                'solve_time': solve_time
            }
    
    def analyze_solution(self, result):
        """分析方案"""
        solver = result['solver']
        u_r = result['variables']['u_r']
        u_pkr = result['variables']['u_pkr']
        x_prt = result['variables']['x_prt']
        
        total_classes = sum(1 for (k, r) in u_r if solver.Value(u_r[(k, r)]) == 1)
        
        class_sizes = []
        for k in self.subjects:
            for r in range(1, self.config['max_classes_per_subject'] + 1):
                if solver.Value(u_r[(k, r)]) == 1:
                    size = sum(self.packages[p]['人数'] for p in self.package_names if solver.Value(u_pkr[(p, k, r)]) == 1)
                    class_sizes.append(size)
        
        split_count = 0
        split_details = []
        for p in self.package_names:
            for group_name, group_slots in self.SLOT_GROUPS.items():
                subjects_in_group = set()
                for t in group_slots:
                    for k in self.subjects:
                        for r in range(1, self.config['max_classes_per_subject'] + 1):
                            if solver.Value(x_prt[(p, k, r, t)]) == 1:
                                subjects_in_group.add(k)
                
                if len(subjects_in_group) >= 2:
                    split_count += 1
                    split_details.append(f"{p}-{group_name}: {', '.join(sorted(subjects_in_group))}")
        
        return {
            'total_classes': total_classes,
            'avg_size': round(sum(class_sizes) / len(class_sizes), 1) if class_sizes else 0,
            'min_size': min(class_sizes) if class_sizes else 0,
            'max_size': max(class_sizes) if class_sizes else 0,
            'split_count': split_count,
            'split_details': split_details
        }
    
    def extract_timetable(self, result):
        """
        提取课表数据（精确格子映射版）
        1. 修复：不再假设课程连续。通过 relative_slots 传递精确的时间槽索引。
           解决 "物理(2h)" 被 "化学(1h)" 覆盖或错位的问题。
        2. 排序：按科目名称聚类排序。
        """
        solver = result['solver']
        u_r = result['variables']['u_r']
        y_rt = result['variables']['y_rt']
        u_pkr = result['variables']['u_pkr']
        
        class_name_map = {} 
        for k in self.subjects:
            active_classes = []
            for r in range(1, self.config['max_classes_per_subject'] + 1):
                if solver.Value(u_r[(k, r)]) == 1:
                    students = [p for p in self.package_names if solver.Value(u_pkr[(p, k, r)]) == 1]
                    size = sum(self.packages[p]['人数'] for p in students)
                    active_classes.append({'r': r, 'size': size})
            active_classes.sort(key=lambda x: (-x['size'], x['r']))
            
            if len(active_classes) > 1:
                for index, item in enumerate(active_classes):
                    class_name_map[(k, item['r'])] = f"班{chr(65 + index)}"
            else:
                for item in active_classes:
                    class_name_map[(k, item['r'])] = "班"

        class_details = []
        for k in self.subjects:
            for r in range(1, self.config['max_classes_per_subject'] + 1):
                if solver.Value(u_r[(k, r)]) == 1:
                    students = [p for p in self.package_names if solver.Value(u_pkr[(p, k, r)]) == 1]
                    size = sum(self.packages[p]['人数'] for p in students)
                    time_slots = [t for t in self.TIME_SLOTS_1H if solver.Value(y_rt[(k, r, t)]) == 1]
                    slot_groups_used = defaultdict(list)
                    for t in time_slots:
                        slot_groups_used[self.SLOT_TO_GROUP[t]].append(t)
                    slot_str = ', '.join([f"{g}({len(slots)}h)" for g, slots in sorted(slot_groups_used.items(), key=lambda x: natural_sort_key(x[0]))])
                    class_details.append({
                        '科目': k,
                        '班级': class_name_map.get((k, r), f'班{r}'),
                        '人数': size,
                        '时段': slot_str,
                        '学生配套': ', '.join(sorted(students, key=natural_sort_key))
                    })
        class_details.sort(key=lambda x: (x['科目'], x['班级']))

        slot_schedule_data = []
        
        for group_name in sorted(self.SLOT_GROUPS.keys(), key=natural_sort_key):
            group_slots = self.SLOT_GROUPS[group_name]
            group_start_time = min(group_slots)
            group_slots_set = set(group_slots)
            
            # 3.1 收集碎片
            fragments = []
            for k in self.subjects:
                for r in range(1, self.config['max_classes_per_subject'] + 1):
                    active_slots = [t for t in group_slots if solver.Value(y_rt[(k, r, t)]) == 1]
                    actual_hours = len(active_slots)
                    if actual_hours == 0: continue
                    students = [p for p in self.package_names if solver.Value(u_pkr[(p, k, r)]) == 1]
                    if not students: continue
                    
                    fragments.append({
                        'subject': f"{k}",
                        'duration_str': f"{actual_hours}h",
                        'class_name': class_name_map.get((k, r), f'班{r}'),
                        'packages_str': ', '.join(sorted(students, key=natural_sort_key)),
                        'raw_packages': students,
                        'size': sum(self.packages[p]['人数'] for p in students),
                        'raw_hours': actual_hours,
                        'active_slots': set(active_slots),
                        'start_time': min(active_slots),
                        'is_gap': False
                    })
            
            # 3.2 贪心拼图
            fragments.sort(key=lambda x: -x['size'])
            visual_rows = []
            for frag in fragments:
                placed = False
                for row in visual_rows:
                    conflict = False
                    for existing in row:
                        if not frag['active_slots'].isdisjoint(existing['active_slots']):
                            conflict = True; break
                    if not conflict:
                        row.append(frag); placed = True; break
                if not placed: visual_rows.append([frag])
            
            # 3.3 填空 & 格式化
            for row_items in visual_rows:
                occupied_slots = set()
                for item in row_items: occupied_slots.update(item['active_slots'])
                missing_slots = sorted(list(group_slots_set - occupied_slots))
                if missing_slots:
                    import itertools
                    for _, g in itertools.groupby(enumerate(missing_slots), lambda ix: ix[0] - ix[1]):
                        gap_group = list(map(lambda ix: ix[1], g))
                        row_items.append({
                            'subject': '0',
                            'duration_str': f"{len(gap_group)}h",
                            'class_name': '-',
                            'packages_str': '-',
                            'raw_packages': [],
                            'size': 0,
                            'raw_hours': 0,
                            'active_slots': set(gap_group),
                            'start_time': min(gap_group),
                            'is_gap': True
                        })
                
                row_items.sort(key=lambda x: x['start_time'])
                
                merged_items_str = []
                for i in row_items:
                    if i['is_gap']:
                        item_str = f"{i['subject']}({i['duration_str']})"
                    else:
                        cls_short = i['class_name'].replace('班', '') 
                        if cls_short:
                            item_str = f"{i['subject']} {cls_short}({i['duration_str']})"
                        else:
                            item_str = f"{i['subject']}({i['duration_str']})"
                    merged_items_str.append(item_str)
                
                merged_info = " + ".join(merged_items_str)
                merged_packages = " + ".join([i['packages_str'] for i in row_items])
                
                unique_pkgs = set()
                for i in row_items:
                    for p in i['raw_packages']: unique_pkgs.add(p)
                unique_count = sum(self.packages[p]['人数'] for p in unique_pkgs)
                
                # UI Display Items
                display_list = []
                for idx, item in enumerate(row_items):
                    ui_class = item['class_name'].replace('班', '')
                    
                    relative_slots = [t - group_start_time for t in item['active_slots']]
                    
                    display_list.append({
                        'seq': idx + 1,
                        'subject': item['subject'],
                        'duration': item['duration_str'],
                        'class': ui_class,
                        'color_seed': item['subject'] if not item['is_gap'] else 'gap',
                        'is_gap': item['is_gap'],
                        'packages_str': item['packages_str'],
                        'relative_slots': relative_slots # <--- 传递精确的格子索引
                    })

                slot_schedule_data.append({
                    '时段': group_name,
                    '时长': f"{sum(i['raw_hours'] for i in row_items)}h",
                    '科目 & 班级': merged_info,
                    '人数': unique_count,
                    '涉及配套': merged_packages,
                    'display_items': display_list,
                    'sort_key_subject': row_items[0]['subject'] if row_items else ""
                })
        
        # 排序：先按科目名，再按时段
        slot_schedule_data.sort(key=lambda x: (natural_sort_key(x['时段']), x['sort_key_subject']))

        return class_details, slot_schedule_data
        
def check_data_feasibility(packages, subject_hours, config):
    """
    全方位预检：人数容量 + 教师资源 + 学生负荷
    """
    issues = []
    
    # 基础配置获取
    min_s = config['min_class_size']
    max_s = config['max_class_size']
    max_k = config['max_classes_per_subject']
    num_slots = config['num_slots']
    

    total_system_hours = (num_slots - 1) * 2 + 3
    
    concurrency_limit = config.get('default_concurrency', 1) 

    enrollment = calculate_subject_enrollment(packages)
    
    for subject, total_students in enrollment.items():
        is_capacity_feasible = False
        min_classes_needed = 0
        
        for r in range(1, max_k + 1):
            if r * min_s <= total_students <= r * max_s:
                is_capacity_feasible = True
                min_classes_needed = r # 记录最少需要开几个班
                break
        
        if not is_capacity_feasible:
            issues.append({
                'type': '人数容量',
                'subject': subject,
                'detail': f"人数({total_students})无法被分配到1-{max_k}个班级中(班额{min_s}-{max_s})。",
                'suggestion': "调整班额限制或最大班数。"
            })
            continue # 人数都排不下，后面不用算了


        hours_per_class = subject_hours.get(subject, 0)

        total_slots_needed = min_classes_needed * hours_per_class

        max_slots_available = total_system_hours * concurrency_limit
        
        if total_slots_needed > max_slots_available:
            issues.append({
                'type': '资源瓶颈',
                'subject': subject,
                'detail': f"需要排 {min_classes_needed} 个班 × {hours_per_class}小时 = {total_slots_needed} 小时，但系统只有 {total_system_hours} 小时可用(并发={concurrency_limit})。",
                'suggestion': f"增加【时段组数量】，或者允许【{subject}】多班并发上课。"
            })


    for pkg_name, pkg_data in packages.items():
        # 计算该配套的总课时
        total_pkg_hours = sum(pkg_data['科目'].values())
        
        if total_pkg_hours > total_system_hours:
            issues.append({
                'type': '学生负荷',
                'subject': pkg_name, # 这里借用字段显示配套名
                'detail': f"该配套学生需要上课 {total_pkg_hours} 小时，但排课总时长只有 {total_system_hours} 小时。",
                'suggestion': "增加【时段组数量】或减少该配套科目。"
            })

    return issues
    
def calculate_smart_defaults(packages, subject_hours, default_concurrency=1):
    """
    计算【真·理论底线】参数
    去掉人为的 40 人保底，完全基于数学除法
    """
    import math
    
    enrollment = calculate_subject_enrollment(packages)
    if not enrollment:
        return {}

    min_student_count = min(enrollment.values())
    if min_student_count > 16:
        raw_min = math.floor(min_student_count / current_max_classes)
        suggested_min_size = max(1, raw_min - 2)
    else:

        suggested_min_size = max(1, min_student_count - 3)
    max_student_count = max(enrollment.values())
    
    assumed_max_classes = 1

    raw_max_size = math.ceil(max_student_count / assumed_max_classes)
    
    suggested_max_size = raw_max_size + 3

    # 3. 时段组数量 (保持不变)
    max_subject_hours_needed = 0
    for subj, hours in subject_hours.items():
        est_classes = math.ceil(enrollment[subj] / suggested_max_size)
        slots_needed = (est_classes * hours) / default_concurrency
        if slots_needed > max_subject_hours_needed:
            max_subject_hours_needed = slots_needed
            
    max_package_hours = 0
    for pkg in packages.values():
        total_h = sum(pkg['科目'].values())
        if total_h > max_package_hours:
            max_package_hours = total_h
            
    hard_limit_hours = max(max_subject_hours_needed, max_package_hours)
    
    if hard_limit_hours <= 3:
        suggested_slots = 1
    else:
        suggested_slots = math.ceil((hard_limit_hours - 3) / 2) + 1
        
    return {
        'min_class_size': int(suggested_min_size),
        'max_class_size': int(suggested_max_size),
        'num_slots': int(max(suggested_slots, 8)) 
    } 
    
def on_max_classes_change():
    """
    当【每科目最大班数】改变时触发：
    自动重新计算并更新【最大班额】的建议值
    """
    if 'packages' not in st.session_state and 'solutions' not in st.session_state:
        return
        

    current_max_classes = st.session_state.get('param_max_classes', 3)
    
    enrollment = calculate_subject_enrollment(st.session_state['packages'])
    if not enrollment:
        return
    max_student_count = max(enrollment.values())
    
    import math
    raw_new_size = math.ceil(max_student_count / current_max_classes)
    suggested_new_size = raw_new_size + 3
    

    st.session_state['param_max_size'] = int(suggested_new_size)
    
    # 5. (可选) 给个提示
    st.toast(f"已根据 {current_max_classes} 个班重新计算，最大班额调整为 {suggested_new_size} 人", icon="🔄")


import pickle
import os

HISTORY_FILE = "schedule_history.pkl"

def load_history_from_disk():
    """读取本地历史记录"""
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        with open(HISTORY_FILE, 'rb') as f:
            return pickle.load(f)
    except:
        return []

def preprocess_and_split_packages(original_packages, max_class_size=24):
    """自动拆分超大配套 (命名优化版)"""
    import math
    new_packages = {}
    split_log = []
    
    for pkg_name, pkg_data in original_packages.items():
        count = pkg_data['人数']
        
        if count <= max_class_size:
            new_packages[pkg_name] = pkg_data
            continue
            
        # === 需要拆分 ===
        num_chunks = math.ceil(count / max_class_size)
        base_size = count // num_chunks
        remainder = count % num_chunks
        
        chunks = []
        for i in range(num_chunks):
            size = base_size + (1 if i < remainder else 0)
            chunks.append(size)
            
        log_entry = {'original': pkg_name, 'total': count, 'parts': []}
        
        for idx, size in enumerate(chunks):
            # 🔥 修改点：使用 A, B, C... 后缀
            suffix = chr(65 + idx) # 0->A, 1->B
            sub_name = f"{pkg_name}_{suffix}" 
            
            new_packages[sub_name] = {
                '人数': size,
                '科目': pkg_data['科目'] 
            }
            log_entry['parts'].append(f"{sub_name}({size}人)")
            
        split_log.append(log_entry)
        
    return new_packages, split_log
    
def analyze_teacher_needs(slot_schedule):
    """
    分析师资需求：精确到每小时计算最大并发数
    """
    from collections import defaultdict
    global_slot_usage = defaultdict(lambda: defaultdict(int))

    teacher_needs = defaultdict(int)

    for slot_data in slot_schedule:
        group_time_usage = defaultdict(lambda: defaultdict(int))
        
        for item in slot_data.get('display_items', []):
            if item.get('is_gap', False):
                continue
                
            subj = item['subject']
            # 获取该课占用的相对格子，例如 [0, 1] 表示前两小时
            relative_slots = item.get('relative_slots', [])
            
            # 如果没有 relative_slots (兼容旧逻辑)，尝试解析
            if not relative_slots:
                 try:
                    dur = int(item['duration'].replace('h',''))
                 except: 
                    dur = 1
                 # 这是一个估算，存在风险，最好确保 upstream 传递了 relative_slots
                 relative_slots = range(dur) 

            for t_idx in relative_slots:
                group_time_usage[t_idx][subj] += 1
        
        # 统计该组内，每个科目在任意时刻的最大并发
        for t_idx, subj_counts in group_time_usage.items():
            for subj, count in subj_counts.items():
                if count > teacher_needs[subj]:
                    teacher_needs[subj] = count
                    
    return teacher_needs

# ==============================================================================
# [增强版] 本地存储工具 (自动修正数据格式 + 调试反馈)
# ==============================================================================
import pickle
import os
import datetime

HISTORY_FILE = "schedule_history.pkl"

def clean_data_for_storage(obj):
    """递归将 set 转为 list，确保可以被序列化"""
    if isinstance(obj, set):
        return list(obj)
    elif isinstance(obj, dict):
        return {k: clean_data_for_storage(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_data_for_storage(i) for i in obj]
    else:
        return obj

def save_history_to_disk(current_solutions):
    """
    保存历史记录 (带强制清洗)
    """
    if not current_solutions:
        print("❌ 保存失败：当前没有方案数据")
        return
    
    # 1. 白名单过滤
    KEYS_TO_SAVE = [
        'name', 'status', 'solve_status', 'solve_time', 'icon', 
        'analysis', 'class_details', 'slot_schedule', 'split_log'
    ]
    
    sanitized_solutions = []
    for sol in current_solutions:
        # 提取白名单数据
        safe_sol = {k: sol[k] for k in KEYS_TO_SAVE if k in sol}
        # 🔥 关键：深度清洗，把 set 转为 list，防止 pickle 报错
        safe_sol = clean_data_for_storage(safe_sol)
        sanitized_solutions.append(safe_sol)
    
    # 2. 读取旧历史
    history = []
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'rb') as f:
                history = pickle.load(f)
        except:
            history = []
    
    # 3. 追加新记录
    timestamp = get_malaysia_time().strftime("%Y-%m-%d %H:%M:%S")
    # 避免重复保存完全一样的数据
    if not history or history[-1]['data'] != sanitized_solutions:
        history.append({'time': timestamp, 'data': sanitized_solutions})
        print(f"✅ 历史记录已追加: {timestamp}")
    else:
        print("⚠️ 数据未变，跳过保存")
    
    # 4. 限制数量（保留最近10条）
    if len(history) > 10:
        history = history[-10:]
        
    # 5. 写入
    try:
        with open(HISTORY_FILE, 'wb') as f:
            pickle.dump(history, f)
        # 在界面上显示个小绿标，证明运行到了这里
        st.toast(f"已保存到本地记录 ({timestamp})", icon="💾")
    except Exception as e:
        st.error(f"❌ 保存文件失败: {str(e)}")
        print(f"❌ 保存错误: {e}")
        
# main design
def main():
    # 初始化 session_state 用于保存方案
    if 'saved_solutions' not in st.session_state:
        # 从磁盘加载已保存的方案
        st.session_state['saved_solutions'] = load_saved_solutions_from_disk()
    
    st.markdown('<div class="main-header">📚 智能排课求解器</div>', unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; color: #666;">走班制排课搜索系统</p>', unsafe_allow_html=True)
    # ... (st.set_page_config 之后) ...

    # [新增] 注入 JS 拦截刷新/关闭事件
    # 这会在用户试图刷新页面时弹窗警告
    import streamlit.components.v1 as components
    components.html(
        """
        <script>
        window.addEventListener('beforeunload', function (e) {
            e.preventDefault();
            e.returnValue = '';
        });
        </script>
        """,
        height=0,
    )
    # 侧边栏
    with st.sidebar:
        st.header("⚙️ 系统配置")
        st.markdown("---")
        st.subheader("📜 历史记录")
        st.caption("💡 **点击下方按钮加载**：无需上传配套即可查看历史记录")
        st.caption(f"📊 保留最近 10 次记录 | 当前: {len(load_history_from_disk())} 条")
        
        history_records = load_history_from_disk()
        
        if not history_records:
            st.caption("暂无历史记录")
        else:
            # 倒序显示，最近的在最上面
            for idx, record in enumerate(reversed(history_records)):
                # idx=0 是最后一场, idx=1 是倒数第二场
                btn_label = f"📂 加载: {record['time']} (共{len(record['data'])}个方案)"
                
                # 使用唯一的 key 防止冲突
                if st.button(btn_label, key=f"hist_btn_{idx}", use_container_width=True):
                    # 加载历史记录
                    st.session_state['solutions'] = record['data']
                    st.session_state['from_history'] = True  # 标记来自历史记录
                    
                    # 添加调试信息
                    st.session_state['debug_loaded'] = True
                    st.session_state['debug_solutions_count'] = len(record['data'])
                    
                    # 立即显示加载结果（在侧边栏）
                    st.success(f"✅ 数据已加载！共 {len(record['data'])} 个方案")
                    st.info("👉 请向下滚动主界面查看方案详情")
                    
                    st.toast(f"✅ 已加载 {record['time']} 的排课结果！共{len(record['data'])}个方案", icon="🎉")
                    # 不使用 st.rerun()，让 Streamlit 自然重新运行
            
            if st.button("🗑️ 清空历史", type="secondary", key="clear_hist"):
                if os.path.exists(HISTORY_FILE):
                    os.remove(HISTORY_FILE)
                    st.toast("✅ 历史记录已清空", icon="🗑️")
                    st.rerun()
        
        st.markdown("---")
        st.subheader("💾 已保存的方案")
        st.caption("⚠️ **注意**：在 Streamlit Cloud 上，保存的方案会在应用重启后丢失。如需永久保存，请下载Excel文件。")
        
        if st.session_state['saved_solutions']:
            st.caption(f"✅ 当前共 {len(st.session_state['saved_solutions'])} 个方案")
            for save_name in list(st.session_state['saved_solutions'].keys()):
                saved_data = st.session_state['saved_solutions'][save_name]
                with st.expander(f"📁 {save_name}"):
                    st.caption(f"{saved_data['original_name']}")
                    st.caption(f"{saved_data['timestamp']}")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("📥", key=f"view_{save_name}"):
                            st.session_state['solutions'] = [saved_data['solution']]
                            st.session_state['from_saved'] = True  # 标记来自保存方案
                            st.toast(f"✅ 已加载方案: {save_name}", icon="📁")
                            st.rerun()
                    with col2:
                        if st.button("🗑️", key=f"del_{save_name}"):
                            delete_saved_solution(save_name)
                            st.toast(f"🗑️ 已删除方案: {save_name}", icon="🗑️")
                            st.rerun()
        else:
            st.caption("暂无保存")
        
        st.markdown("---")
        
        # 调试信息
        with st.expander("🔍 系统状态", expanded=False):
            # 显示调试信息
            if st.session_state.get('debug_loaded', False):
                st.success("✅ 上次加载成功")
                st.caption(f"加载的方案数: {st.session_state.get('debug_solutions_count', 0)}")
                st.session_state['debug_loaded'] = False  # 显示后清除
            
            if 'solutions' in st.session_state:
                st.success(f"✅ 当前加载: {len(st.session_state['solutions'])} 个方案")
                for i, sol in enumerate(st.session_state['solutions']):
                    name = sol.get('name', '未知')
                    has_details = '✅' if 'class_details' in sol else '❌'
                    has_schedule = '✅' if 'slot_schedule' in sol else '❌'
                    st.caption(f"{i+1}. {name} (详情:{has_details} 时段:{has_schedule})")
            else:
                st.info("暂无加载的方案")
            
            st.caption(f"📜 历史记录: {len(load_history_from_disk())} 条")
            st.caption(f"💾 保存方案: {len(st.session_state['saved_solutions'])} 个")
        
        st.markdown("---")
        st.subheader("📁 数据导入")
        
        # 下载模板功能
        st.markdown("##### 📥 下载数据模板")
        st.markdown("""
        <div style="padding: 0.5rem; border-radius: 0.3rem; margin-bottom: 0.5rem; font-size: 0.85rem;">
        💡 首次使用？下载示例模板了解数据格式
        </div>
        """, unsafe_allow_html=True)
        
        # 创建示例CSV数据
        template_data = """配套,科目,人数,总学点
P12,"会计学（4）,经济（4）,商业（3）,历史（4）,AI应用（2）,AI编程（2）",5,19
P13,"物理（6）,经济（4）,历史（4）,地理（4）,AI应用（2）",6,20
P14,"物理（6）,会计学（4）,经济（4）,商业（3）,AI应用（2）,AI编程（2）",4,21
P15,"生物（4）,化学（5）,物理（6）,会计学（4）,AI应用（2）",9,21
P16,"生物（4）,化学（5）,物理（6）,商业（3）,AI应用（2）",3,20
P17,"生物（4）,化学（5）,会计学（4）,地理（4）,AI应用（2）,AI编程（2）",8,21
P18,"生物（4）,化学（5）,经济（4）,历史（4）,AI应用（2）,AI编程（2）",11,21
P19,"物理（6）,经济（4）,商业（3）,历史（4）,AI应用（2）,AI编程（2）",7,21
P20,"物理（6）,生物（4）,化学（5）,经济（4）,AI应用（2）",10,21
P21,"物理（6）,生物（4）,化学（5）,地理（4）,AI应用（2）",2,21
P22,"生物（4）,化学（5）,经济（4）,地理（4）,AI应用（2）,AI编程（2）",12,21"""
        
        # 下载按钮
        col1, col2 = st.columns([1, 1])
        with col1:
            st.download_button(
                label="📄 CSV模板",
                data=template_data.encode('utf-8-sig'),  # 使用BOM确保Excel正确识别UTF-8
                file_name="排课数据模板.csv",
                mime="text/csv",
                help="下载CSV格式的示例模板",
                use_container_width=True
            )
        with col2:
            # 创建Excel格式的模板
            template_df = pd.DataFrame([
                {'配套': 'P12', '科目': '会计学（4）,经济（4）,商业（3）,历史（4）,AI应用（2）,AI编程（2）', '人数': 5, '总学点': 19},
                {'配套': 'P13', '科目': '物理（6）,经济（4）,历史（4）,地理（4）,AI应用（2）', '人数': 6, '总学点': 20},
                {'配套': 'P14', '科目': '物理（6）,会计学（4）,经济（4）,商业（3）,AI应用（2）,AI编程（2）', '人数': 4, '总学点': 21},
                {'配套': 'P15', '科目': '生物（4）,化学（5）,物理（6）,会计学（4）,AI应用（2）', '人数': 9, '总学点': 21},
            ])
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                template_df.to_excel(writer, index=False, sheet_name='配套数据')
            
            st.download_button(
                label="📊 Excel模板",
                data=excel_buffer.getvalue(),
                file_name="排课数据模板.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="下载Excel格式的示例模板",
                use_container_width=True
            )
        
        st.markdown("---")
        
        # 文件上传
        st.markdown("##### 📤 上传数据文件")
        uploaded_file = st.file_uploader(
            "选择文件",
            type=['xlsx', 'xls', 'csv'],
            help="支持Excel和CSV格式，需包含'配套'、'科目'、'人数'列",
            label_visibility="collapsed"
        )
        
        # ... (在 st.file_uploader 之后) ...
    
        if uploaded_file:
            is_new_file = False
            if 'last_uploaded_file' not in st.session_state or st.session_state['last_uploaded_file'] != uploaded_file.name:
                is_new_file = True
                st.session_state['last_uploaded_file'] = uploaded_file.name
    
            with st.spinner("正在解析文件..."):
                packages, subject_hours, max_hours = parse_uploaded_file(uploaded_file)
            
            if packages and subject_hours:
                # 存入 session
                st.session_state['packages'] = packages
                st.session_state['subject_hours'] = subject_hours
                st.session_state['max_total_hours'] = max_hours
    
                # === 🔥 核心修改：如果是新文件，自动计算并填充参数 ===
                if is_new_file:
                    defaults = calculate_smart_defaults(packages, subject_hours)
                    
                    # 直接更新 session_state，这会改变下方输入框的默认值
                    st.session_state['param_min_size'] = defaults['min_class_size']
                    st.session_state['param_max_size'] = defaults['max_class_size']
                    st.session_state['param_num_slots'] = defaults['num_slots']
                    
                    st.toast(f"已根据数据自动调整：最小班额{defaults['min_class_size']}人, 最大{defaults['max_class_size']}人, 时段{defaults['num_slots']}组", icon="🪄")
                # ====================================================
        
        st.markdown("---")
        
        st.subheader("🔧 求解参数")
        
        # 1. 最小班额
        if 'param_min_size' not in st.session_state:
            st.session_state['param_min_size'] = 5 # 初始默认值
            
        min_class_size = st.number_input(
            "最小班额", 
            min_value=1, max_value=100, 
            key="param_min_size", # <--- 绑定到 Session State
            step=1
        )

        # 2. 最大班额
        if 'param_max_size' not in st.session_state:
            st.session_state['param_max_size'] = 60
            
        st.info("💡提示: 已自动计算最低有解要求，若无解可以手动把最大班额+1，如果要极致减少开班数可以把班额上限尽量放大")    
        max_class_size = st.number_input(
            "最大班额", 
            min_value=1, max_value=200, 
            key="param_max_size", # <--- 绑定到 Session State
            step=1
        )
        
        # 3. 每科目最大班数
        max_classes_per_subject = st.number_input(
            "每科目最大班数", 
            min_value=1, max_value=10, 
            step=1,
            key="param_max_classes",      # <--- 1. 绑定 Key
            on_change=on_max_classes_change # <--- 2. 绑定刚才写的回调函数
        )
        
        # 4. 时段组数量
        if 'param_num_slots' not in st.session_state:
             # 原来的逻辑：根据 max_hours 推荐，或者默认 10
             if 'max_total_hours' in st.session_state:
                 st.session_state['param_num_slots'] = calculate_recommended_slots(st.session_state['max_total_hours'])
             else:
                 st.session_state['param_num_slots'] = 10

        num_slots = st.number_input(
            "时段组数量", 
            min_value=1, max_value=30, 
            key="param_num_slots", # <--- 绑定到 Session State
            step=1,
            help="系统已根据总学时自动计算推荐值"
        )
        
        # 智能推荐时段组数
        if 'max_total_hours' in st.session_state:
            max_hours = st.session_state['max_total_hours']
            recommended_slots = calculate_recommended_slots(max_hours)
            total_capacity = (recommended_slots - 1) * 2 + 3
            
            st.markdown(f"""
            <div style="padding: 1rem; border-radius: 0.5rem; border-left: 4px solid #2196f3; margin: 1rem 0;">
                <strong>📊 智能分析</strong><br>
                • 最大总课时：<strong>{max_hours}小时</strong><br>
                • 推荐时段组数：<strong>{recommended_slots}组</strong> (总容量{total_capacity}小时)<br>
                • 说明：{recommended_slots-1}组×2小时 + 1组×3小时 = {total_capacity}小时
            </div>
            """, unsafe_allow_html=True)
            
            default_slots = recommended_slots
        else:
            default_slots = 10
            st.info("💡 上传数据后将自动推荐时段组数")
        
        num_slots = st.number_input(
            "时段组数量", 
            min_value=1, 
            max_value=20, 
            value=default_slots, 
            step=1,
            help="系统会根据数据自动推荐，也可手动调整。最后一个时段组为3小时，其余为2小时"
        )
        
        st.info("💡提示: 增加求解时间可得更优解，上限600")
        solver_timeout = st.number_input("求解超时(秒)", min_value=10, max_value=600, value=120, step=10)
        
        st.markdown("---")
        
        st.subheader("🔀 时段分割")
        allow_slot_split = st.checkbox("允许时段分割", value=True,
                                      help="允许一个时段内上不同科目的课")
        if allow_slot_split:
            slot_split_penalty = st.slider("分割惩罚系数", min_value=0, max_value=5000, value=1000, step=100,
                                          help="越大越不愿意分割")
        else:
            slot_split_penalty = 0
        
        st.markdown("---")

        st.markdown("##### ✂️ 方案D配置")
        scheme_d_limit = st.number_input(
            "方案D自动拆分上限", 
            min_value=10, max_value=100, value=24, step=1,
            help="当配套人数超过此数值时，自动拆分为多个小配套（方案D专用）"
        )
        
        # === ✨ 新增：方案选择器 ✨ ===
        st.markdown("---")
        st.subheader("🎯 方案选择")
        
        # 定义选项映射
        SCHEME_OPTIONS = [
            "方案A: 最少开班 (传统模式)",
            "方案B: 全局均衡 (避免拥挤)",
            "方案C: 精品小班 (强控30人)",
            "方案D: 自动拆分 (解决超大班)"
        ]
        
        # 多选框，默认全选
        selected_schemes_ui = st.multiselect(
            "勾选需要运行的方案",
            options=SCHEME_OPTIONS,
            default=SCHEME_OPTIONS,
            help="取消勾选不需要的方案可节省计算时间"
        )
        # ============================
        
        st.markdown("---")
        
        st.subheader("🔒 强制开班")
        if 'subject_hours' in st.session_state:
            forced_class_count = {}
            for subject in st.session_state['subject_hours'].keys():
                count = st.number_input(f"{subject}", min_value=0, max_value=10, value=0, key=f"forced_{subject}")
                if count > 0:
                    forced_class_count[subject] = count
        else:
            forced_class_count = {}
            st.info("请先上传数据文件")
    
    # 主内容区
    # 如果既没有上传文件，也没有加载历史记录，显示使用说明
    if 'packages' not in st.session_state and 'solutions' not in st.session_state:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("""
        ### 智能排课搜索器
        
        **使用步骤：**
        1. 📁 在左侧上传配套数据文件（Excel或CSV格式）
        2. ⚙️ 调整求解参数（可选）
        3. 🚀 点击"开始求解"按钮
        4. 📊 查看并下载结果
        
        **或者：**
        - 📂 点击左侧栏的"历史记录"加载之前的求解结果（无需上传文件）
        
        **数据格式要求：**
        - 必须包含列：`配套`、`科目`、`人数`
        - 科目格式：`会计(6),历史(4),地理(4)` 或 `会计（6）,历史（4）`

    
    
        **功能：**
        - 🎯 自动生成多个优化方案
        - 🔀 支持时段分割（一个时段上不同科目）
        - 👨‍🏫 教师资源约束（同科目不同班不冲突）
        - 📊 时段总表（查看每个时段的全局安排）
        - ⏰ 灵活课时
        """)
        st.markdown('</div>', unsafe_allow_html=True)
        return
    
    # 显示数据概览
    if 'packages' in st.session_state:
        st.markdown('<div class="sub-header">📊 数据概览</div>', unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("配套数量", len(st.session_state['packages']))
        with col2:
            st.metric("科目数量", len(st.session_state['subject_hours']))
        with col3:
            total_students = sum(p['人数'] for p in st.session_state['packages'].values())
            st.metric("学生总数", total_students)
    
        # 配套详情
        with st.expander("查看配套详情"):
            df_packages = []
            for name, data in st.session_state['packages'].items():
                subjects_str = ', '.join([f"{k}({v}h)" for k, v in data['科目'].items()])
                df_packages.append({
                    '配套': name,
                    '人数': data['人数'],
                    '科目': subjects_str
                })
            st.dataframe(pd.DataFrame(df_packages), use_container_width=True)
        
        # 科目选修统计
        with st.expander("查看科目选修统计"):
            enrollment = calculate_subject_enrollment(st.session_state['packages'])
            df_enrollment = pd.DataFrame([
                {'科目': k, '课时': st.session_state['subject_hours'][k], '选修人数': enrollment[k]}
                for k in sorted(enrollment.keys(), key=lambda x: enrollment[x], reverse=True)
            ])
            st.dataframe(df_enrollment, use_container_width=True)
    
    st.markdown("---")
    
    # ========== 最近求解记录（自动显示）==========
    st.markdown('<div class="sub-header">📋 最近求解记录</div>', unsafe_allow_html=True)
    
    # 加载最近2次历史记录
    recent_history = load_history_from_disk()
    
    if recent_history:
        # 只显示最近2次
        display_records = list(reversed(recent_history))[:2]
        
        if len(display_records) > 0:
            st.caption(f"自动显示最近 {len(display_records)} 次求解结果（无需上传配套）")
            
            for idx, record in enumerate(display_records):
                with st.expander(f"📊 {record['time']} - 共 {len(record['data'])} 个方案", expanded=(idx==0)):
                    # 显示方案对比表格
                    comparison_data = []
                    for sol in record['data']:
                        if 'analysis' in sol:
                            analysis = sol['analysis']
                            comparison_data.append({
                                '方案': sol['name'],
                                '开班数': analysis['total_classes'],
                                '平均班额': f"{analysis['avg_size']}人",
                                '班额范围': f"{analysis['min_size']}-{analysis['max_size']}人",
                                '时段分割': analysis['split_count'],
                                '状态': sol.get('icon', '✅')
                            })
                        else:
                            comparison_data.append({
                                '方案': sol.get('name', '未知'),
                                '开班数': '-',
                                '平均班额': '-',
                                '班额范围': '-',
                                '时段分割': '-',
                                '状态': sol.get('icon', '❌')
                            })
                    
                    if comparison_data:
                        df_comparison = pd.DataFrame(comparison_data)
                        st.dataframe(df_comparison, use_container_width=True)
                    
                    # 加载按钮
                    col1, col2 = st.columns([4, 1])
                    with col2:
                        if st.button("📥 加载到主界面", key=f"load_recent_{idx}", use_container_width=True):
                            st.session_state['solutions'] = record['data']
                            st.session_state['from_history'] = True
                            st.toast(f"✅ 已加载 {record['time']} 的方案", icon="🎉")
                            st.rerun()
    else:
        st.info("暂无历史记录。完成第一次求解后，这里会自动显示最近的结果。")
    
    st.markdown("---")

    current_config = {
        'min_class_size': min_class_size,
        'max_class_size': max_class_size,
        'max_classes_per_subject': max_classes_per_subject,
        'num_slots': num_slots,  # <--- 补上这一行！
        # 如果你之前加了并发功能，记得把这一行也加上，否则会报 'default_concurrency' 错误
        'default_concurrency': st.session_state.get('default_concurrency', 1) if 'default_concurrency' in st.session_state else 1
        # 或者如果你把并发输入框定义为了变量 default_concurrency，则写: 'default_concurrency': default_concurrency
    }

    # --- 插入点：实时预检 ---
    if 'packages' in st.session_state:
        feasibility_issues = check_data_feasibility(
            st.session_state['packages'], 
            st.session_state['subject_hours'], 
            current_config
        )
        
        if feasibility_issues:
            st.markdown('<div class="error-box">', unsafe_allow_html=True)
            st.error(f"⚠️ 检测到 {len(feasibility_issues)} 个科目存在数学逻辑冲突（必无解）：")
            
            for issue in feasibility_issues:
                error_type = issue.get('type', '错误')
                detail = issue.get('detail', issue.get('reason', '未知原因'))
                
                st.markdown(f"""
                **❌ [{error_type}] {issue['subject']}**: {detail}
                * <small style="color: #666;">建议: {issue['suggestion']}</small>
                """, unsafe_allow_html=True)
            
            st.warning("💡 请调整上方的【最小班额】、【最大班额】或【每科目最大班数】，直到此错误框消失。")
            st.markdown('</div>', unsafe_allow_html=True)
            
            # 可选：如果存在致命错误，禁用求解按钮
            disable_solve = True
        else:
            st.success("✅ 数据校验通过：所有科目的总人数均在合法区间内。")
            disable_solve = False
    else:
        disable_solve = True

    st.markdown("---")
    st.markdown('<div class="sub-header">🚀 开始求解</div>', unsafe_allow_html=True)
    
    if st.button("🎯 生成排课方案", type="primary", use_container_width=True, disabled=disable_solve):
        config = {
            'min_class_size': min_class_size,
            'max_class_size': max_class_size,
            'max_classes_per_subject': max_classes_per_subject,
            'num_slots': num_slots,
            'allow_slot_split': allow_slot_split,
            'slot_split_penalty': slot_split_penalty,
            'forced_class_count': forced_class_count
        }
        
        solver_instance = ScheduleSolver(
            st.session_state['packages'],
            st.session_state['subject_hours'],
            config
        )
        
        # === ✨ 修改：根据选择动态生成配置 ✨ ===
        if not selected_schemes_ui:
            st.error("❌ 请至少选择一个方案！")
            return # 停止运行

        solution_configs = []
        
        # 按顺序判断，确保运行顺序 A->B->C->D
        if "方案A: 最少开班 (传统模式)" in selected_schemes_ui:
            solution_configs.append({'type': 'min_classes', 'name': '方案A：最少开班'})
            
        if "方案B: 全局均衡 (避免拥挤)" in selected_schemes_ui:
            solution_configs.append({'type': 'balanced', 'name': '方案B：全局均衡'})
            
        if "方案C: 精品小班 (强控30人)" in selected_schemes_ui:
            solution_configs.append({'type': 'subject_balanced', 'name': '方案C：精品小班(上限30人)'})
            
        if "方案D: 自动拆分 (解决超大班)" in selected_schemes_ui:
            solution_configs.append({'type': 'auto_split', 'name': f'方案D：自动拆分(上限{scheme_d_limit}人)'})
        # ======================================
        
        # 进度条初始化
        progress_container = st.container()
        with progress_container:
            progress_bar = st.progress(0)
            col1, col2 = st.columns([3, 1])
            with col1:
                status_text = st.empty()
            with col2:
                percentage_text = st.empty()

        solutions = []
        total_steps = len(solution_configs) * 3 
        current_step = 0
        
        import math # 确保导入

        for i, sol_config in enumerate(solution_configs):

            run_config = config.copy()
            current_packages = st.session_state['packages']
            split_info = None # 用于记录拆分日志

            if sol_config['type'] == 'auto_split':
                # 1. 执行拆分
                new_pkgs, logs = preprocess_and_split_packages(
                    st.session_state['packages'], 
                    max_class_size=scheme_d_limit
                )
                current_packages = new_pkgs # 切换为拆分后的数据
                split_info = logs
                
                if logs:
                    status_text.markdown(f"✂️ **{sol_config['name']}** - 已拆分 {len(logs)} 个超大配套...")
                    time.sleep(0.5)
                
                # 2. 借用方案C的内核，但使用方案D的参数
                sol_config['type'] = 'subject_balanced' 
                
                # 3. 强制覆盖参数 (保证有解)
                enrollment = calculate_subject_enrollment(current_packages)
                max_students = max(enrollment.values()) if enrollment else 0
                import math
                theoretical_needed = math.ceil(max_students / scheme_d_limit)
                
                # 班数给够
                run_config['max_classes_per_subject'] = int(theoretical_needed + 2)
                run_config['min_class_size'] = 1
                run_config['dynamic_max_limit'] = scheme_d_limit
                run_config['forced_class_count'] = {}
                run_config['enable_concurrency'] = True                
                run_config['relax_hard_lock'] = True

            elif sol_config['type'] == 'subject_balanced':
                enrollment = calculate_subject_enrollment(st.session_state['packages'])
                max_students = max(enrollment.values()) if enrollment else 0
                scheme_c_limit = 30
                theoretical_needed = math.ceil(max_students / scheme_c_limit)
                run_config['max_classes_per_subject'] = int(theoretical_needed + 2)
                
                run_config['min_class_size'] = 1
                status_text.markdown(f"🔓 **{sol_config['name']}** - 已启用独立规则 (忽略全局参数，自动计算班数...)")
                time.sleep(0.5)

            solver_instance = ScheduleSolver(
                current_packages, 
                st.session_state['subject_hours'],
                run_config
            )

            current_step += 1
            progress = current_step / total_steps
            progress_bar.progress(progress)
            status_text.markdown(f"🔄 **{sol_config['name']}** - 准备数据...")
            percentage_text.markdown(f"**{int(progress * 100)}%**")
            
            current_step += 1
            progress = current_step / total_steps
            progress_bar.progress(progress)
            status_text.markdown(f"🏗️ **{sol_config['name']}** - 构建数学模型...")
            percentage_text.markdown(f"**{int(progress * 100)}%**")
            
            model, variables = solver_instance.build_model(sol_config['type'])
            
            current_step += 1
            progress = current_step / total_steps
            progress_bar.progress(progress)
            status_text.markdown(f"⚙️ **{sol_config['name']}** - 启动求解引擎...")
            percentage_text.markdown(f"**{int(progress * 100)}%**")
            
            result = solver_instance.solve(
                model, 
                variables, 
                solver_timeout,
                status_placeholder=status_text,
                scheme_name=sol_config['name']
            )
            
            if result['status'] == 'success':
                result['name'] = sol_config['name']
                result['analysis'] = solver_instance.analyze_solution(result)
                result['class_details'], result['slot_schedule'] = solver_instance.extract_timetable(result)
                solutions.append(result)
                status_text.markdown(f"✅ **{sol_config['name']}** - 求解完成")
            else:
                status_text.markdown(f"❌ **{sol_config['name']}** - 求解失败")
                time.sleep(1)
        
        progress_bar.progress(1.0)
        percentage_text.markdown("**100%**")
        status_text.markdown("🎉 **所有方案求解完成！**")
        time.sleep(0.5)
        
        progress_bar.empty()
        status_text.empty()
        percentage_text.empty()
        
        if not solutions:
            st.markdown('<div class="error-box">', unsafe_allow_html=True)
            st.error("❌ 所有方案均无解！")
            st.markdown("""
            **可能原因：**
            - 时段数量不足
            - 班额限制过严
            - 强制开班数设置不合理
            
            **建议解决方案：**
            1. 增加时段组数量
            2. 放宽班额上限
            3. 取消强制开班限制
            4. 启用时段分割功能
            """)
            st.markdown('</div>', unsafe_allow_html=True)
            return
        
        st.session_state['solutions'] = solutions
        
        # Show Solution
        st.markdown('<div class="success-box">', unsafe_allow_html=True)
        st.success(f"✅ 成功生成 {len(solutions)} 个方案！")
        st.markdown('</div>', unsafe_allow_html=True)
        save_history_to_disk(solutions)
    
    if 'solutions' in st.session_state:
        st.markdown("---")
        
        # 🔥 调试信息：确认进入显示方案的代码块
        st.success(f"🔍 调试：检测到 {len(st.session_state['solutions'])} 个方案待显示")
        
        # 如果是从历史记录或保存的方案加载的，显示提示
        if st.session_state.get('from_history', False):
            st.info("📂 当前显示的是从历史记录加载的方案（点击左侧栏加载按钮后显示）")
            st.caption("💡 不需要上传配套数据即可查看")
            st.session_state['from_history'] = False  # 显示后清除标记
        elif st.session_state.get('from_saved', False):
            st.info("📁 当前显示的是从已保存方案加载的内容")
            st.caption("💡 不需要上传配套数据即可查看")
            st.session_state['from_saved'] = False  # 显示后清除标记
        
        st.markdown('<div class="sub-header">📊 方案对比</div>', unsafe_allow_html=True)
        
        # 显示方案数量
        st.caption(f"共 {len(st.session_state['solutions'])} 个方案")
        
        comparison_data = []
        for sol in st.session_state['solutions']:
            # 检查方案是否成功
            if 'analysis' in sol:
                analysis = sol['analysis']
                comparison_data.append({
                    '方案': sol['name'],
                    '开班数': analysis['total_classes'],
                    '平均班额': f"{analysis['avg_size']}人",
                    '班额范围': f"{analysis['min_size']}-{analysis['max_size']}人",
                    '时段分割次数': analysis['split_count'],
                    '求解时间': f"{sol.get('solve_time', 0):.1f}秒",
                    '状态': sol.get('icon', '✅')
                })
            else:
                # 失败的方案
                comparison_data.append({
                    '方案': sol.get('name', '未知方案'),
                    '开班数': '-',
                    '平均班额': '-',
                    '班额范围': '-',
                    '时段分割次数': '-',
                    '求解时间': f"{sol.get('solve_time', 0):.1f}秒",
                    '状态': sol.get('icon', '❌')
                })
        
        if comparison_data:
            df_comparison = pd.DataFrame(comparison_data)
            st.dataframe(df_comparison, use_container_width=True)
        else:
            st.info("没有可显示的方案数据")
        
        for sol in st.session_state['solutions']:
            # 检查方案数据完整性
            has_details = 'class_details' in sol
            has_schedule = 'slot_schedule' in sol
            has_analysis = 'analysis' in sol
            
            # 如果数据不完整，显示警告
            if not (has_details and has_schedule):
                with st.expander(f"⚠️ {sol.get('name', '未知方案')} - 数据不完整", expanded=False):
                    st.warning("此方案的数据不完整，无法显示详细信息")
                    st.caption("可能原因：")
                    st.caption("- 方案求解失败")
                    st.caption("- 历史记录数据格式较旧")
                    st.caption("- 数据保存时出现问题")
                    if not has_details:
                        st.caption("❌ 缺少: class_details (开班详情)")
                    if not has_schedule:
                        st.caption("❌ 缺少: slot_schedule (时段总表)")
                    if not has_analysis:
                        st.caption("❌ 缺少: analysis (统计分析)")
                continue
                
            with st.expander(f"📋 {sol['name']} - 详细结果"):

                if 'split_log' in sol:
                    st.info("✂️ **自动拆分方案**：以下大配套已被拆分为 A/B 班")
                    split_data = []
                    for log in sol['split_log']:
                        split_data.append({
                            '原配套': log['original'],
                            '总人数': log['total'],
                            '拆分结果': ' + '.join(log['parts']), # 例如 P1_A(12人) + P1_B(13人)
                            '班数': len(log['parts'])
                        })
                    st.dataframe(pd.DataFrame(split_data), use_container_width=True)
                
                if sol['name'].startswith('方案D') and sol['status'] == 'success':
                    st.markdown("##### 👨‍🏫 师资与开班统计")
                    teacher_needs = analyze_teacher_needs(sol['slot_schedule'])
                    
                    # 整理数据
                    stats_data = []
                    total_classes_map = defaultdict(int)
                    for item in sol['class_details']:
                        total_classes_map[item['科目']] += 1
                    
                    for subj in sorted(total_classes_map.keys()):
                        stats_data.append({
                            '科目': subj,
                            '总开班数': total_classes_map[subj], 
                            '所需老师(并发数)': teacher_needs.get(subj, 1), 
                            '单班平均': f"{round(sum(c['人数'] for c in sol['class_details'] if c['科目']==subj)/total_classes_map[subj], 1)}人"
                        })
                    
                    st.dataframe(pd.DataFrame(stats_data), use_container_width=True)
    
                st.markdown("---")

                tab1, tab2, tab3 = st.tabs(["开班详情", "时段总表", "数据导出"])
                
                with tab1:
                    df_class = pd.DataFrame(sol['class_details'])
                    st.dataframe(df_class, use_container_width=True)
                    
                    if sol['analysis']['split_count'] > 0:
                        st.markdown('<div class="warning-box">', unsafe_allow_html=True)
                        st.warning(f"⚠️ 检测到 {sol['analysis']['split_count']} 处时段分割")
                        for detail in sol['analysis']['split_details']:
                            st.text(f"  • {detail}")
                        st.markdown('</div>', unsafe_allow_html=True)
                
                with tab2:
                    st.markdown("### 🕐 时段总表")
                    
                    schedule_data = sol['slot_schedule']
                    if not schedule_data:
                        st.info("暂无数据")
                    else:
                        
                        schedule_data = sol['slot_schedule']
                    if not schedule_data:
                        st.info("暂无数据")
                    else:
                        # 🔥🔥🔥 [修复版] CSS：完美适配 Light/Dark Mode 🔥🔥🔥
                        # 🔥🔥🔥 [修复版] 修复了最后一行看不清的问题 🔥🔥🔥
                        table_css = """
                        <style>
                            /* 全局表格样式 */
                            .schedule-table { 
                                width: 100%; 
                                border-collapse: collapse; 
                                font-family: sans-serif; 
                                margin-bottom: 1rem; 
                                font-size: 14px; 
                                table-layout: fixed; 
                                background-color: var(--background-color); 
                                color: var(--text-color);
                            }
                            
                            /* 表头样式 */
                            .schedule-table th { 
                                background-color: var(--secondary-background-color); 
                                color: var(--text-color); 
                                padding: 10px 6px; 
                                text-align: center; 
                                border-bottom: 2px solid var(--primary-color); 
                                border-top: 1px solid rgba(128,128,128,0.2); 
                            }
                            
                            /* 单元格样式 */
                            .schedule-table td { 
                                padding: 6px; 
                                text-align: left; 
                                border-right: 1px solid rgba(128,128,128,0.1); 
                                border-bottom: 1px solid rgba(128,128,128,0.1); 
                                vertical-align: middle; 
                                color: var(--text-color); 
                            }
                            
                            /* 左侧时段列 */
                            .col-slot { 
                                width: 50px; 
                                font-weight: 800; 
                                color: var(--primary-color); 
                                background-color: var(--secondary-background-color); 
                                border-right: 2px solid rgba(128,128,128,0.2) !important; 
                                text-align: center !important;
                            }
                            
                            /* 分组底部分割线 (关键修复) */
                            /* 之前这里有 opacity: 0.1 导致整行看不清，已删除 */
                            .group-border-bottom { 
                                border-bottom: 2px solid rgba(128, 128, 128, 0.3) !important; 
                            }
                            
                            /* 辅助列 */
                            .col-duration { width: 40px; text-align: center !important; opacity: 0.8; }
                            .col-count { width: 40px; text-align: center !important; font-weight: bold; }
                            .col-pkg { width: 20%; font-size: 0.85rem; text-align: center !important; opacity: 0.7; }
                            
                            /* 卡片容器 */
                            .timeline-container { display: flex; align-items: center; flex-wrap: wrap; gap: 4px; }
                            
                            /* 课程卡片 */
                            .timeline-card { 
                                background-color: var(--secondary-background-color); 
                                border: 1px solid rgba(128,128,128,0.2); 
                                border-radius: 4px; 
                                padding: 3px 6px; 
                                display: flex; 
                                flex-direction: column; 
                                min-width: 80px; 
                                box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                            }
                            
                            /* 空档卡片 */
                            .card-gap {
                                background-color: transparent !important;
                                border: 1px dashed rgba(128,128,128,0.3) !important;
                                opacity: 0.6;
                            }
                            
                            .card-header { display: flex; align-items: center; margin-bottom: 2px; }
                            
                            .seq-badge { 
                                background-color: var(--primary-color); 
                                color: white; 
                                font-size: 0.7rem; 
                                font-weight: bold; 
                                width: 16px; height: 16px; 
                                border-radius: 50%; 
                                display: flex; align-items: center; justify-content: center; 
                                margin-right: 5px; 
                            }
                            
                            .subject-name { 
                                font-weight: 700; 
                                font-size: 0.9rem; 
                                color: var(--text-color); 
                            }
                            
                            .card-footer { 
                                display: flex; 
                                justify-content: space-between; 
                                font-size: 0.75rem; 
                                color: var(--text-color); 
                                opacity: 0.7; 
                            }
                            
                            .arrow-icon { 
                                color: var(--text-color); 
                                opacity: 0.3; 
                                font-size: 1rem; 
                                margin: 0 2px; 
                            }
                        </style>
                        """
                        
                        # 重新生成 HTML 行 (逻辑保持不变，只需修改 CSS 类名引用的部分)
                        html_rows = []
                        from itertools import groupby
                        schedule_data.sort(key=lambda x: (natural_sort_key(x['时段']), x.get('sort_key_subject', '')))
                        
                        for slot_name, items in groupby(schedule_data, key=lambda x: x['时段']):
                            group_items = list(items)
                            row_count = len(group_items)
                            for i, item in enumerate(group_items):
                                border_class = "group-border-bottom" if i == row_count - 1 else ""
                                row_html = f"<tr class='{border_class}'>"
                                if i == 0:
                                    row_html += f"<td class='col-slot' rowspan='{row_count}'>{item['时段']}</td>"
                                    row_html += f"<td class='col-duration' rowspan='{row_count}'>{item['时长']}</td>"
                                
                                flow_html = '<div class="timeline-container">'
                                display_items = item.get('display_items', [])
                                for idx, d_item in enumerate(display_items):
                                    # 使用 CSS 类而不是内联样式
                                    card_class = "timeline-card card-gap" if d_item['is_gap'] else "timeline-card"
                                    badge_style = "opacity: 0.2;" if d_item['is_gap'] else "" # 仅对 gap 做透明度处理，颜色走 CSS
                                    
                                    card = f"""<div class="{card_class}"><div class="card-header"><span class="seq-badge" style="{badge_style}">{d_item['seq']}</span><span class="subject-name">{d_item['subject']}</span></div><div class="card-footer"><span>{d_item['class']}</span><span>{d_item['duration']}</span></div></div>"""
                                    flow_html += card
                                    if idx < len(display_items) - 1: flow_html += '<div class="arrow-icon">➜</div>'
                                flow_html += '</div>'
                                
                                row_html += f"<td>{flow_html}</td>"
                                row_html += f"<td class='col-count'>{item['人数']}</td>"
                                
                                pkg_slots = ["-", "-", "-"]
                                
                                for d_item in display_items:

                                    relative_slots = d_item.get('relative_slots', [])
                                    
                                    if not relative_slots and 'start_offset' in d_item:
                                         try:
                                            dur = int(d_item['duration'].replace('h',''))
                                         except: dur = 1
                                         start = d_item['start_offset']
                                         relative_slots = range(start, start + dur)

                                    pkg_str = d_item.get('packages_str', '-')
                                    if not pkg_str or d_item.get('is_gap', False): 
                                        pkg_str = "-"
                                    

                                    for slot_idx in relative_slots:
                                        if 0 <= slot_idx < 3:
                                            pkg_slots[slot_idx] = pkg_str
                                
                                for grid_idx in range(3):
                                    row_html += f"<td class='col-pkg'>{pkg_slots[grid_idx]}</td>"
                                
                                row_html += "</tr>"
                                html_rows.append(row_html)
                        
                        full_html = f"""
                        {table_css}
                        <table class="schedule-table">
                            <thead>
                                <tr>
                                    <th class="col-slot">时段</th>
                                    <th class="col-duration">长</th>
                                    <th>课程流程</th>
                                    <th class="col-count">数</th>
                                    <th class="col-pkg">第 1 小时</th>
                                    <th class="col-pkg">第 2 小时</th>
                                    <th class="col-pkg">第 3 小时</th>
                                </tr>
                            </thead>
                            <tbody>{''.join(html_rows)}</tbody>
                        </table>
                        """
                        st.markdown(full_html, unsafe_allow_html=True)

                    # Show result
                    st.markdown("### 📊 统计信息")
                    df_slot = pd.DataFrame(schedule_data)
                    cols_to_drop = ['display_items', 'sort_key_subject']
                    df_slot_export = df_slot.drop(columns=[c for c in cols_to_drop if c in df_slot.columns])
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("总时段数", df_slot['时段'].nunique() if not df_slot.empty else 0)
                    with col2:
                        st.metric("总条目数", len(df_slot))
                    with col3:
                        unique = df_slot['时段'].nunique() if not df_slot.empty else 0
                        avg = len(df_slot) / unique if unique > 0 else 0
                        st.metric("平均每时段条目", f"{avg:.1f}")
                # Export              
                with tab3:
                    # 导出为Excel
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # 准备数据源
                        raw_class_data = sol['class_details']
                        raw_slot_data = sol['slot_schedule']
                        
                        df_class = pd.DataFrame(raw_class_data)
                        
                        # [通用函数] 定义合并逻辑
                        def format_subject_class_col(row):
                            suffix = row['班级'].replace('班', '')
                            if suffix:
                                return f"{row['科目']} {suffix}"
                            else:
                                return row['科目']

                        # =========================================================
                        # 1. 处理 "开班详情" Sheet
                        # =========================================================
                        df_class = df_class.sort_values(by=['科目', '班级'])
                        df_class['科目 & 班级'] = df_class.apply(format_subject_class_col, axis=1)
                        df_class_export = df_class[['科目 & 班级', '人数', '时段', '学生配套']]
                        df_class_export.to_excel(writer, sheet_name='开班详情', index=False)
                        
                        
                        # =========================================================
                        # 2. 处理 "时段总表" Sheet
                        # =========================================================
                        df_slot = pd.DataFrame(raw_slot_data)
                        
                        # 准备 3 个新列
                        p1_list, p2_list, p3_list = [], [], []
                        
                        for item in raw_slot_data:
                            current_pkg_slots = ["-", "-", "-"]
                            d_items = item.get('display_items', [])
                            
                            if isinstance(d_items, list):
                                for sub_item in d_items:
                                    pkg_str = sub_item.get('packages_str', '-')
                                    if not pkg_str or sub_item.get('is_gap', False):
                                        pkg_str = "-"
                                    
                                    # 获取精确槽位
                                    rel_slots = sub_item.get('relative_slots', [])
                                    # Fallback
                                    if not rel_slots and 'start_offset' in sub_item:
                                        try: dur = int(sub_item['duration'].replace('h',''))
                                        except: dur = 1
                                        start = sub_item['start_offset']
                                        rel_slots = range(start, start + dur)
                                        
                                    for idx in rel_slots:
                                        if 0 <= idx < 3:
                                            current_pkg_slots[idx] = pkg_str
                            
                            p1_list.append(current_pkg_slots[0])
                            p2_list.append(current_pkg_slots[1])
                            p3_list.append(current_pkg_slots[2])
                        
                        # 添加新列
                        df_slot['配套 (第1小时)'] = p1_list
                        df_slot['配套 (第2小时)'] = p2_list
                        df_slot['配套 (第3小时)'] = p3_list
                        
                        # 剔除无关列
                        drops = ['display_items', 'sort_key_subject', '涉及配套']
                        df_slot = df_slot.drop(columns=[c for c in drops if c in df_slot.columns])
                        
                        # 调整列顺序
                        base_cols = [c for c in df_slot.columns if '配套' not in c]
                        new_cols = ['配套 (第1小时)', '配套 (第2小时)', '配套 (第3小时)']
                        df_slot = df_slot[base_cols + new_cols]
                        
                        # 写入 Excel
                        df_slot.to_excel(writer, sheet_name='时段总表', index=False)
                        
                        # =========================================================
                        # [核心修复] Excel 样式处理：先合并，后画线
                        # =========================================================
                        from openpyxl.styles import Alignment, Border, Side
                        
                        ws_slot = writer.sheets['时段总表']
                        col_pkg_start = 5 
                        
                        # 样式定义
                        thick_border = Border(bottom=Side(style='thick', color='000000'))
                        thin_border = Border(bottom=Side(style='thin', color='D3D3D3'))
                        center_align = Alignment(horizontal='center', vertical='center')
                        
                        max_row = len(df_slot) + 1 
                        slot_merge_start = 2
                        
                        for r_idx in range(2, max_row + 2):
                            # --- A. 配套列横向合并逻辑 ---
                            cell1 = ws_slot.cell(row=r_idx, column=col_pkg_start)
                            cell2 = ws_slot.cell(row=r_idx, column=col_pkg_start+1)
                            cell3 = ws_slot.cell(row=r_idx, column=col_pkg_start+2)
                            
                            val1, val2, val3 = cell1.value, cell2.value, cell3.value
                            
                            if val1 == val2 == val3 and val1 != '-':
                                ws_slot.merge_cells(start_row=r_idx, start_column=col_pkg_start, end_row=r_idx, end_column=col_pkg_start+2)
                                cell1.alignment = center_align
                            elif val1 == val2 and val1 != '-':
                                ws_slot.merge_cells(start_row=r_idx, start_column=col_pkg_start, end_row=r_idx, end_column=col_pkg_start+1)
                                cell1.alignment = center_align
                                cell3.alignment = center_align
                            elif val2 == val3 and val2 != '-':
                                ws_slot.merge_cells(start_row=r_idx, start_column=col_pkg_start+1, end_row=r_idx, end_column=col_pkg_start+2)
                                cell2.alignment = center_align
                                cell1.alignment = center_align
                            else:
                                cell1.alignment = center_align
                                cell2.alignment = center_align
                                cell3.alignment = center_align
                            
                            # --- B. 分组判断逻辑 ---
                            current_slot = ws_slot.cell(row=r_idx, column=1).value
                            next_slot = None
                            if r_idx < max_row + 1:
                                next_slot = ws_slot.cell(row=r_idx+1, column=1).value
                            
                            # 如果到达分组边界
                            if current_slot != next_slot:
                                # 1. [先] 纵向合并时段列 (S1...) 和 时长列 (2h...)
                                # 即使 r_idx == slot_merge_start (单行)，合并也是安全的
                                ws_slot.merge_cells(start_row=slot_merge_start, start_column=1, end_row=r_idx, end_column=1)
                                ws_slot.merge_cells(start_row=slot_merge_start, start_column=2, end_row=r_idx, end_column=2)
                                
                                # 设置居中对齐 (针对合并后的左上角单元格)
                                ws_slot.cell(row=slot_merge_start, column=1).alignment = center_align
                                ws_slot.cell(row=slot_merge_start, column=2).alignment = center_align
                                
                                # 2. [后] 画粗底边 (Outline) - 修复 Bug
                                # 即使第1、2列已经合并了，我们依然要给 row=r_idx (该组最后一行) 的所有单元格设置底边框。
                                # Excel 会根据合并区域最底部单元格的边框来渲染整体边框。
                                for c_idx in range(1, 8):
                                    cell = ws_slot.cell(row=r_idx, column=c_idx)
                                    cell.border = thick_border
                                
                                # 更新下一组起始行
                                slot_merge_start = r_idx + 1
                            else:
                                # 组内画浅色线
                                for c_idx in range(1, 8):
                                    ws_slot.cell(row=r_idx, column=c_idx).border = thin_border

                        
                        # =========================================================
                        # 3. 处理 "所有班级及涉及的配套" Sheet
                        # =========================================================
                        df_overview = df_class_export[['科目 & 班级', '学生配套', '人数']].copy()
                        df_overview.columns = ['科目 SUBJECT', '配套 PACKAGE', '人数']
                        df_overview.to_excel(writer, sheet_name='导入', index=False)
                        
                        
                        # =========================================================
                        # 4. 自动调整列宽
                        # =========================================================
                        workbook = writer.book
                        for sheet_name in writer.sheets:
                            worksheet = writer.sheets[sheet_name]
                            if sheet_name == '时段总表':
                                df_to_measure = df_slot
                            elif sheet_name == '导入':
                                df_to_measure = df_overview
                            else:
                                df_to_measure = df_class_export
                                
                            for idx, col in enumerate(df_to_measure.columns):
                                max_len = max(
                                    len(str(col)),
                                    df_to_measure[col].astype(str).str.len().max() if not df_to_measure[col].empty else 0
                                )
                                adjusted_width = min(max_len + 4, 60)
                                worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width
                    
                    st.download_button(
                        label="📥 下载Excel文件",
                        data=output.getvalue(),
                        file_name=f"{sol['name'].replace('：', '_')}_排课结果.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.markdown("---")
                    st.markdown("#### 💾 保存方案")
                    
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        save_name = st.text_input(
                            "输入存储名称",
                            placeholder="例如：2024秋季排课_最终版",
                            key=f"save_name_{sol['name']}"
                        )
                    with col2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("💾 保存", key=f"save_{sol['name']}"):
                            if save_name:
                                if save_name in st.session_state['saved_solutions']:
                                    st.warning(f"⚠️ '{save_name}' 已存在")
                                else:
                                    save_solution_to_storage(sol, save_name)
                                    st.success(f"✅ 已保存")
                                    time.sleep(1)
                                    st.rerun()
                            else:
                                st.error("❌ 请输入名称")
if __name__ == "__main__":
    main()
